"""IO utilities for reading, writing, and managing files and folders.

This module provides utility functions for loading various file types,
including Jupyter notebooks, Excel workbooks, JSON, CSV, and Python code files.
It also includes notebook manipulation utilities.
"""

import json
from pathlib import Path
import pandas as pd
import nbformat
from openpyxl import load_workbook
import nbformat
import ast
from pathlib import Path
from nbformat import validate, ValidationError

import nbformat
from nbformat import validate, ValidationError
import nbformat
from nbformat import validate, ValidationError
from uuid import uuid4


def safe_load_notebook(path: Path) -> nbformat.NotebookNode:
    """Safely load a Jupyter notebook and ensure cell IDs exist.
    
    Reads a Jupyter notebook and automatically adds unique IDs to cells
    that are missing them. This prevents MissingIDFieldWarning in nbformat
    versions >=5.1.4. The file is not modified on disk.
    
    Parameters
    ----------
    path : Path
        Path to the Jupyter notebook file (.ipynb).
    
    Returns
    -------
    nbformat.NotebookNode
        The loaded and validated notebook object with all cells having IDs.
    
    Raises
    ------
    RuntimeError
        If the notebook cannot be loaded or validated.
    
    Examples
    --------
    >>> from pathlib import Path
    >>> nb = safe_load_notebook(Path("notebook.ipynb"))
    >>> print(f"Loaded {len(nb.cells)} cells")
    
    Notes
    -----
    This function modifies the notebook structure in memory only. It does not
    write changes back to the file system.
    """
    try:
        nb = nbformat.read(path, as_version=4)

        # Manually ensure IDs exist for all cells
        modified = False
        for cell in nb.cells:
            if "id" not in cell:
                cell["id"] = str(uuid4())
                modified = True

        # Try validating the notebook
        try:
            validate(nb)
        except ValidationError:
            # Upgrade schema if older version
            nb = nbformat.v4.upgrade(nb)
            validate(nb)

        # (Optional) Log in debug mode if any IDs were added
        if modified:
            print(f"[safe_load_notebook] Added missing IDs in memory for {path.name}")

        return nb

    except Exception as e:
        raise RuntimeError(f"Unable to load notebook {path}: {e}")


def normalize_notebook(path=None, inplace: bool = True) -> nbformat.NotebookNode:
    """Normalize a Jupyter notebook by ensuring unique cell IDs.
    
    Ensures every cell in a Jupyter notebook has a unique 'id' field.
    This prevents MissingIDFieldWarning in nbformat >=5.1.4.
    
    Parameters
    ----------
    path : str or Path
        Path to the notebook file (.ipynb).
    inplace : bool, default=True
        If True, overwrites the notebook file with normalized version.
        If False, returns the normalized object without modifying the file.
    
    Returns
    -------
    nbformat.NotebookNode
        The normalized notebook object with validated cell IDs.
    
    Examples
    --------
    >>> from pathlib import Path
    >>> nb = normalize_notebook("notebook.ipynb", inplace=False)
    >>> print(f"Normalized {len(nb.cells)} cells")
    
    See Also
    --------
    safe_load_notebook : Load notebook without modifying the file.
    """
    path = Path(path)
    nb = nbformat.read(path, as_version=4)

    # Add missing IDs if needed
    nbformat.validate(nb)  # may warn, but we fix below
    normalized_nb = nbformat.v4.upgrade(nb)
    nbformat.v4.validate_cell_ids(normalized_nb)

    if inplace:
        nbformat.write(normalized_nb, path)
        return normalized_nb

    return normalized_nb


def load_notebook(path: Path) -> nbformat.NotebookNode:
    """Load a Jupyter notebook file.
    
    Parameters
    ----------
    path : Path
        Path to the notebook file (.ipynb).
    
    Returns
    -------
    nbformat.NotebookNode
        The loaded notebook object.
    
    Examples
    --------
    >>> from pathlib import Path
    >>> nb = load_notebook(Path("notebook.ipynb"))
    """
    return nbformat.read(path, as_version=4)


def load_excel(path: Path):
    """Load an Excel workbook.
    
    Parameters
    ----------
    path : Path
        Path to the Excel file (.xlsx or .xlsm).
    
    Returns
    -------
    openpyxl.Workbook
        The loaded Excel workbook object with formulas preserved.
    
    Examples
    --------
    >>> from pathlib import Path
    >>> wb = load_excel(Path("data.xlsx"))
    >>> print(wb.sheetnames)
    """
    return load_workbook(path, data_only=False)


def load_json(path: Path) -> dict:
    """Load a JSON file.
    
    Parameters
    ----------
    path : Path
        Path to the JSON file.
    
    Returns
    -------
    dict
        The parsed JSON data as a dictionary.
    
    Examples
    --------
    >>> from pathlib import Path
    >>> data = load_json(Path("config.json"))
    """
    with open(path, "r", encoding="utf8") as f:
        return json.load(f)


def load_csv(path: Path) -> pd.DataFrame:
    """Load a CSV file into a pandas DataFrame.
    
    Parameters
    ----------
    path : Path
        Path to the CSV file.
    
    Returns
    -------
    pandas.DataFrame
        The loaded CSV data.
    
    Examples
    --------
    >>> from pathlib import Path
    >>> df = load_csv(Path("data.csv"))
    >>> print(df.head())
    """
    return pd.read_csv(path)


def load_raw_code(path: Path) -> str:
    """Load raw Python code from a file.
    
    Parameters
    ----------
    path : Path
        Path to the Python file (.py).
    
    Returns
    -------
    str
        The raw Python code as a string.
    
    Examples
    --------
    >>> from pathlib import Path
    >>> code = load_raw_code(Path("script.py"))
    >>> print(len(code))
    """
    return Path(path).read_text(encoding="utf8")


import ast
import nbformat
from pathlib import Path


def generate_student_notebook(instructor_path: str | Path, output_path: str | Path):
    """Create a student version of an instructor Jupyter notebook.
    
    Generates a student notebook by:
    - Replacing all function bodies with 'pass' statements
    - Removing code cells containing assertions
    - Removing inline assert statements
    - Keeping all Markdown cells and overall structure
    
    Parameters
    ----------
    instructor_path : str or Path
        Path to the instructor's notebook (.ipynb).
    output_path : str or Path
        Destination path for the student version (.ipynb).
    
    Raises
    ------
    FileNotFoundError
        If the instructor notebook does not exist.
    
    Examples
    --------
    >>> generate_student_notebook(
    ...     "instructor_solution.ipynb",
    ...     "student_assignment.ipynb"
    ... )
    ✅ Student notebook generated at: student_assignment.ipynb
    
    Notes
    -----
    This function uses AST parsing to intelligently identify and stub out
    function definitions while preserving imports, assignments, and
    markdown content.
    """
    instructor_path = Path(instructor_path)
    output_path = Path(output_path)

    if not instructor_path.exists():
        raise FileNotFoundError(f"Notebook not found: {instructor_path}")

    nb = nbformat.read(instructor_path, as_version=4)
    new_cells = []

    for cell in nb.cells:
        # --- Keep markdown cells as is ---
        if cell.cell_type == "markdown":
            new_cells.append(cell)
            continue

        if cell.cell_type != "code":
            # Just in case other cell types appear
            new_cells.append(cell)
            continue

        src = cell.source or ""
        stripped_lines = [l.strip() for l in src.splitlines() if l.strip()]

        # --- NEW: drop any code cell that contains an assert line ---
        if any(line.startswith("assert ") for line in stripped_lines):
            # This is considered a test/assert cell → remove entirely
            continue

        # From here on, we are in a non-assert code cell (e.g., function definitions)
        try:
            tree = ast.parse(src)
            new_body: list[ast.stmt] = []

            for node in tree.body:
                # Replace function implementations with stubs
                if isinstance(node, ast.FunctionDef):
                    node.body = [ast.Pass()]
                    new_body.append(node)

                # Keep top-level imports, assignments, and expressions
                elif isinstance(node, (ast.Import, ast.ImportFrom, ast.Assign, ast.Expr)):
                    new_body.append(node)

                # Ignore anything else (for safety), or you can choose to keep it

            # Rebuild the code cell
            new_module = ast.Module(body=new_body, type_ignores=[])
            new_source = ast.unparse(new_module).strip()

            if new_source:
                cell.source = new_source
                new_cells.append(cell)

        except Exception:
            # Fallback: keep non-assert lines only, and still replace functions if desired
            cleaned_lines = []
            for line in src.splitlines():
                stripped = line.strip()
                # drop any inline assert as an extra safeguard
                if stripped.startswith("assert "):
                    continue
                cleaned_lines.append(line)

            new_src = "\n".join(cleaned_lines).strip()
            if new_src:
                cell.source = new_src
                new_cells.append(cell)

    # --- Rebuild notebook ---
    new_nb = nbformat.v4.new_notebook()
    new_nb.cells = new_cells
    new_nb.metadata = nb.metadata

    output_path.parent.mkdir(parents=True, exist_ok=True)
    nbformat.write(new_nb, output_path)
    print(f"✅ Student notebook generated at: {output_path}")


def remove_notebook_with_line(directory: str | Path, line: str):
    """Remove notebooks containing a specific line of code.
    
    Searches for and deletes all Jupyter notebook files in the specified
    directory that contain a particular line of code in any code cell.
    
    Parameters
    ----------
    directory : str or Path
        The directory to search for Jupyter notebook files.
    line : str
        The exact line of code to search for within notebooks.
    
    Raises
    ------
    ValueError
        If the specified path is not a directory.
    
    Examples
    --------
    >>> remove_notebook_with_line("submissions/", "assert solution == ")
    Deleted notebook: submissions/test_notebook.ipynb
    
    Warnings
    --------
    This function permanently deletes files. Use with caution.
    """
    directory = Path(directory)
    if not directory.is_dir():
        raise ValueError(f"The specified path is not a directory: {directory}")

    for notebook_path in directory.glob("*.ipynb"):
        try:
            nb = nbformat.read(notebook_path, as_version=4)
            for cell in nb.cells:
                if cell.cell_type == "code" and line in cell.source:
                    notebook_path.unlink()  # Delete the notebook file
                    print(f"Deleted notebook: {notebook_path}")
                    break  # No need to check further cells
        except Exception as e:
            print(f"Error processing {notebook_path}: {e}")
